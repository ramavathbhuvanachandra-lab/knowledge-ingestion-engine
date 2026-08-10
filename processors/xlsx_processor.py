from pathlib import Path
import json

from openpyxl import load_workbook


class XLSXProcessor:
    """
    Extract structured content from XLSX documents.

    Responsibilities:

    - Open XLSX workbook.
    - Process every worksheet.
    - Remove completely empty rows/columns.
    - Preserve meaningful table structure.
    - Preserve worksheet boundaries.
    - Convert workbook content into Markdown.
    - Save extraction metadata.
    - Do NOT chunk.
    - Do NOT embed.
    - Do NOT perform retrieval.
    """

    def __init__(
        self,
        output_path: str | Path = "storage/processed_documents",
    ):
        self.output_path = Path(output_path)

    # ------------------------------------------------------------
    # PUBLIC API
    # ------------------------------------------------------------

    def process(
        self,
        xlsx_path: str | Path,
        source_url: str | None = None,
    ) -> Path:
        """
        Extract XLSX content and save as Markdown.

        Returns:
            Path to generated Markdown file.
        """

        xlsx_path = Path(xlsx_path)

        # --------------------------------------------------------
        # VALIDATION
        # --------------------------------------------------------

        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"XLSX file does not exist: {xlsx_path}"
            )

        if not xlsx_path.is_file():
            raise ValueError(
                f"XLSX path is not a file: {xlsx_path}"
            )

        if xlsx_path.suffix.lower() != ".xlsx":
            raise ValueError(
                f"Expected an XLSX file: {xlsx_path}"
            )

        # --------------------------------------------------------
        # READ WORKBOOK
        # --------------------------------------------------------

        workbook = load_workbook(
            filename=xlsx_path,
            read_only=True,
            data_only=True,
        )

        worksheets = workbook.worksheets

        if not worksheets:
            workbook.close()

            raise ValueError(
                f"XLSX contains no worksheets: {xlsx_path}"
            )

        # --------------------------------------------------------
        # PROCESS SHEETS
        # --------------------------------------------------------

        markdown_parts = []

        total_rows = 0
        total_non_empty_rows = 0
        total_cells = 0
        total_non_empty_cells = 0

        sheets_processed = 0

        for worksheet in worksheets:

            sheet_result = self._process_sheet(
                worksheet
            )

            if sheet_result is None:
                continue

            sheets_processed += 1

            markdown_parts.append(
                f"# {worksheet.title}"
            )

            markdown_parts.append("")

            if sheet_result["context"]:
                markdown_parts.extend(
                    sheet_result["context"]
                )

                markdown_parts.append("")

            markdown_parts.append(
                sheet_result["table"]
            )

            markdown_parts.append("")

            total_rows += sheet_result[
                "rows"
            ]

            total_non_empty_rows += (
                sheet_result[
                    "non_empty_rows"
                ]
            )

            total_cells += sheet_result[
                "cells"
            ]

            total_non_empty_cells += (
                sheet_result[
                    "non_empty_cells"
                ]
            )

        workbook.close()

        # --------------------------------------------------------
        # VALIDATE EXTRACTION
        # --------------------------------------------------------

        markdown = "\n".join(
            markdown_parts
        ).strip()

        if not markdown:
            raise ValueError(
                f"No extractable content found in XLSX: "
                f"{xlsx_path}"
            )

        # --------------------------------------------------------
        # OUTPUT DIRECTORY
        # --------------------------------------------------------

        domain = xlsx_path.parent.name

        output_dir = (
            self.output_path
            / domain
        )

        output_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        # --------------------------------------------------------
        # OUTPUT FILE
        # --------------------------------------------------------

        output_file = (
            output_dir
            / f"{xlsx_path.stem}.md"
        )

        output_file.write_text(
            markdown,
            encoding="utf-8",
        )

        # --------------------------------------------------------
        # METADATA
        # --------------------------------------------------------

        metadata = {
            "source_url": source_url,
            "source_file": str(xlsx_path),
            "document_type": "xlsx",
            "sheets": len(worksheets),
            "sheets_processed": sheets_processed,
            "rows": total_rows,
            "non_empty_rows": total_non_empty_rows,
            "cells": total_cells,
            "non_empty_cells": total_non_empty_cells,
            "text_length": len(markdown),
            "extraction_success": True,
        }

        metadata_file = (
            output_dir
            / f"{xlsx_path.stem}.json"
        )

        metadata_file.write_text(
            json.dumps(
                metadata,
                indent=4,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        # --------------------------------------------------------
        # LOGGING
        # --------------------------------------------------------

        print(
            f"Processed XLSX : {xlsx_path}"
        )

        print(
            f"Markdown       : {output_file}"
        )

        print(
            f"Metadata       : {metadata_file}"
        )

        print(
            f"Sheets         : {len(worksheets)}"
        )

        print(
            f"Sheets processed: {sheets_processed}"
        )

        print(
            f"Rows           : {total_rows}"
        )

        print(
            f"Non-empty rows : {total_non_empty_rows}"
        )

        print(
            f"Text length    : {len(markdown)}"
        )

        return output_file

    # ------------------------------------------------------------
    # SHEET PROCESSING
    # ------------------------------------------------------------

    def _process_sheet(
        self,
        worksheet,
    ) -> dict | None:

        rows = []

        total_rows = 0
        non_empty_rows = 0
        total_cells = 0
        non_empty_cells = 0

        for row in worksheet.iter_rows(
            values_only=True
        ):

            total_rows += 1

            values = list(row)

            total_cells += len(values)

            cleaned = [
                self._clean_cell(value)
                for value in values
            ]

            non_empty = [
                value
                for value in cleaned
                if value
            ]

            if not non_empty:
                continue

            non_empty_rows += 1
            non_empty_cells += len(
                non_empty
            )

            rows.append(cleaned)

        if not rows:
            return None

        # --------------------------------------------------------
        # REMOVE EMPTY TRAILING COLUMNS
        # --------------------------------------------------------

        max_columns = max(
            len(row)
            for row in rows
        )

        useful_columns = []

        for column_index in range(
            max_columns
        ):

            has_content = any(
                column_index < len(row)
                and row[column_index]
                for row in rows
            )

            if has_content:
                useful_columns.append(
                    column_index
                )

        normalized_rows = []

        for row in rows:

            normalized_rows.append(
                [
                    row[index]
                    if index < len(row)
                    else ""
                    for index in useful_columns
                ]
            )

        # --------------------------------------------------------
        # FIND TABLE HEADER
        # --------------------------------------------------------

        header_index = self._find_header(
            normalized_rows
        )

        context_rows = (
            normalized_rows[:header_index]
        )

        table_rows = (
            normalized_rows[header_index:]
        )

        if not table_rows:
            return None

        # --------------------------------------------------------
        # BUILD CONTEXT
        # --------------------------------------------------------

        context = []

        for row in context_rows:

            text = " | ".join(
                value
                for value in row
                if value
            )

            if text:
                context.append(
                    text
                )

        # --------------------------------------------------------
        # BUILD TABLE
        # --------------------------------------------------------

        table = self._build_markdown_table(
            table_rows
        )

        return {
            "context": context,
            "table": table,
            "rows": total_rows,
            "non_empty_rows": non_empty_rows,
            "cells": total_cells,
            "non_empty_cells": non_empty_cells,
        }

    # ------------------------------------------------------------
    # HEADER DETECTION
    # ------------------------------------------------------------

    def _find_header(
        self,
        rows: list[list[str]],
    ) -> int:

        for index, row in enumerate(rows):

            non_empty = [
                value
                for value in row
                if value
            ]

            if len(non_empty) >= 2:
                return index

        return 0

    # ------------------------------------------------------------
    # MARKDOWN TABLE
    # ------------------------------------------------------------

    def _build_markdown_table(
        self,
        rows: list[list[str]],
    ) -> str:

        if not rows:
            return ""

        column_count = max(
            len(row)
            for row in rows
        )

        normalized = []

        for row in rows:

            normalized.append(
                [
                    row[index]
                    if index < len(row)
                    else ""
                    for index in range(
                        column_count
                    )
                ]
            )

        header = normalized[0]

        # --------------------------------------------------------
        # MAKE UNIQUE / NON-EMPTY HEADERS
        # --------------------------------------------------------

        headers = []

        for index, value in enumerate(
            header
        ):

            value = value.strip()

            if not value:
                value = f"Column {index + 1}"

            headers.append(
                self._escape_markdown(
                    value
                )
            )

        lines = []

        lines.append(
            "| "
            + " | ".join(headers)
            + " |"
        )

        lines.append(
            "| "
            + " | ".join(
                "---"
                for _ in headers
            )
            + " |"
        )

        for row in normalized[1:]:

            values = [
                self._escape_markdown(
                    value
                )
                for value in row
            ]

            lines.append(
                "| "
                + " | ".join(values)
                + " |"
            )

        return "\n".join(lines)

    # ------------------------------------------------------------
    # CELL CLEANING
    # ------------------------------------------------------------

    def _clean_cell(
        self,
        value,
    ) -> str:

        if value is None:
            return ""

        text = str(value).strip()

        if not text:
            return ""

        text = text.replace(
            "\n",
            " "
        )

        text = text.replace(
            "\r",
            " "
        )

        text = text.replace(
            "|",
            "\\|"
        )

        return " ".join(
            text.split()
        )

    # ------------------------------------------------------------
    # MARKDOWN ESCAPING
    # ------------------------------------------------------------

    def _escape_markdown(
        self,
        value: str,
    ) -> str:

        return value.replace(
            "|",
            "\\|"
        )